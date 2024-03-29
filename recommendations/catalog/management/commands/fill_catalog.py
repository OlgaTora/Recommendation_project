import csv
import os
from more_itertools import unique_everseen
import django
import openpyxl

from django.core.management.base import BaseCommand

from catalog.models import ActivityTypes, ActivityLevel1, ActivityLevel2, ActivityLevel3, Groups, Attends, \
    GroupsCorrect
from users.models import Profile


class Command(BaseCommand):
    help = "Generate catalog from file."

    def handle(self, *args, **kwargs):
        os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'recommendations.settings')
        django.setup()

        workbook = openpyxl.load_workbook('files/dict.xlsx')
        sheet = workbook.active

        def get_set(column: int):
            set_types = set()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                set_types.add(row[column])
            return set_types

        # душа, тело, ум
        set_types = get_set(0)
        for activity in set_types:
            ActivityTypes.objects.create(
                activity_type=activity,
            )

        # направление 1
        lst_types = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            full_row = [row[0], row[1], row[2]]
            lst_types.append(full_row)
        lst_types = list(unique_everseen(lst_types))
        for i in range(len(lst_types)):
            ActivityLevel1.objects.create(
                activity_type=ActivityTypes.objects.get(activity_type=lst_types[i][0]),
                id_level=lst_types[i][1],
                level=lst_types[i][2],
            )

        # направление 2
        lst_types = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            full_row = [row[0], row[1], row[3], row[4]]
            lst_types.append(full_row)
        lst_types = list(unique_everseen(lst_types))
        for i in range(len(lst_types)):
            activity_type = ActivityTypes.objects.get(activity_type=lst_types[i][0])
            level1 = ActivityLevel1.objects.get(id_level=lst_types[i][1], activity_type=activity_type)
            ActivityLevel2.objects.create(
                activity_type=level1,
                id_level=lst_types[i][2],
                level=lst_types[i][3],
            )
        # направление 3
        lst_types = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            lst_types.append(row)
        for i in range(len(lst_types)):
            activity_type = ActivityTypes.objects.get(activity_type=lst_types[i][0])
            level1 = ActivityLevel1.objects.get(id_level=lst_types[i][1], activity_type=activity_type)
            level2 = ActivityLevel2.objects.get(id_level=lst_types[i][3], activity_type=level1)

            ActivityLevel3.objects.create(
                activity_type=level2,
                id_level=lst_types[i][5],
                level=lst_types[i][6],
                descript_level=lst_types[i][7]
            )

        #группы
        with open('files/groups_test.csv', 'r', encoding='utf-8') as groups:
            file_reader = csv.reader(groups, delimiter=',')
            next(file_reader)
            for row in file_reader:
                activity_type = ActivityLevel2.objects.get(level=row[1])
                Groups.objects.create(
                    uniq_id=row[0],
                    level=ActivityLevel3.objects.get(level=row[2], activity_type=activity_type),
                )

        # attends
        with open('files/attends_test.csv', 'r', encoding='utf-8') as attends:
            file_reader = csv.reader(attends, delimiter=',')
            next(file_reader)
            for row in file_reader:
                Attends.objects.create(
                    uniq_id=row[0],
                    group_id=Groups.objects.get(uniq_id=row[1]),
                    user_id=Profile.objects.get(username=row[2]),
                    online=True if row[5] == 'Да' else False,
                    date_attend=row[6],
                    start_time=row[7],
                    end_time=row[8],
                )


        # группы correct
        with open('files/groups_test_corr.csv', 'r', encoding='utf-8') as groups_corr:
            file_reader = csv.reader(groups_corr, delimiter=',')
            next(file_reader)
            for row in file_reader:
                activity_type = ActivityLevel2.objects.get(level=row[2])
                GroupsCorrect.objects.create(
                    group_id=Groups.objects.get(uniq_id=row[1]),
                    level=ActivityLevel3.objects.get(level=row[3], activity_type=activity_type),
                    address=row[4],
                    admin_district=row[5],
                    weekday=row[6],
                    start_date=row[7],
                    end_date=row[8],
                    start_time=row[9],
                    end_time=row[10],
                )
