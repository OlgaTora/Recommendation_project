import csv
import os
from more_itertools import unique_everseen
import django
import openpyxl

from django.core.management.base import BaseCommand
from catalog.models import ActivityTypes, ActivityLevel1, ActivityLevel2, ActivityLevel3, Groups, Attends
from users.models import Profile


class Command(BaseCommand):
    help = "Generate catalog from file."

    def handle(self, *args, **kwargs):
        os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'recommendations.settings')
        django.setup()

        workbook = openpyxl.load_workbook('files/dict.xlsx')
        sheet = workbook.active

        def get_set(column: int):
            set_types = set()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                set_types.add(row[column])
            return set_types

        # душа, тело, ум
        set_types = get_set(0)
        for activity in set_types:
            ActivityTypes.types.create(
                activity_type=activity,
            )

        # направление 1
        lst_types = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            full_row = [row[0], row[1], row[2]]
            lst_types.append(full_row)
        lst_types = list(unique_everseen(lst_types))
        for i in range(len(lst_types)):
            ActivityLevel1.levels.create(
                activity_type=ActivityTypes.types.filter(activity_type=lst_types[i][0]).first(),
                id_level=lst_types[i][1],
                level=lst_types[i][2],
            )

        # направление 2
        lst_types = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            full_row = [row[0], row[1], row[3], row[4]]
            lst_types.append(full_row)
        lst_types = list(unique_everseen(lst_types))
        for i in range(len(lst_types)):
            activity_type = ActivityTypes.types.filter(activity_type=lst_types[i][0]).first()
            level1 = ActivityLevel1.levels.filter(id_level=lst_types[i][1], activity_type=activity_type).first()
            ActivityLevel2.levels.create(
                activity_type=level1,
                id_level=lst_types[i][2],
                level=lst_types[i][3],
            )
        # направление 3
        lst_types = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            lst_types.append(row)
        for i in range(len(lst_types)):
            activity_type = ActivityTypes.types.filter(activity_type=lst_types[i][0]).first()
            level1 = ActivityLevel1.levels.filter(id_level=lst_types[i][1], activity_type=activity_type).first()
            level2 = ActivityLevel2.levels.filter(id_level=lst_types[i][3], activity_type=level1).first()

            ActivityLevel3.levels.create(
                activity_type=level2,
                id_level=lst_types[i][5],
                level=lst_types[i][6],
                descript_level=lst_types[i][7]
            )

        def clean_group_address(address: str):
            """Function for delete replicas in address"""
            address = (address.replace('город', '')
                       .replace('г.', '')
                       .replace('Город', '')
                       .replace('Г', '')
                       .replace('г.о.', ''))
            addresses = address.split('Москва,')
            # убрать пустой список вначале, запятые в конце списка
            addresses = ['город Москва, ' + i.strip().strip(',') for i in addresses if i != ' ' and i != '']
            addresses = ', '.join(list(set(addresses)))
            return addresses

        # группы
        with open('files/groups.csv', 'r', encoding='utf-8') as address_base:
            groups_list = []
            file_reader = csv.reader(address_base, delimiter=',')
            for row in file_reader:
                groups_list.append(row)

        for group in range(1, len(groups_list)):
            address = clean_group_address(groups_list[group][4])
            Groups.groups.create(
                uniq_id=groups_list[group][0],
                level=ActivityLevel3.levels.filter(level=groups_list[group][3]).first(),
                address=address,
                districts=groups_list[group][5].replace('административные округа', 'административный округ'),
                schedule_active=groups_list[group][7],
                schedule_past=groups_list[group][8],
                schedule_plan=groups_list[group][9],
            )
            print(group)

        # attends
        with open('files/attends.csv', 'r', encoding='utf-8') as attends:
            attends_list = []
            file_reader = csv.reader(attends, delimiter=',')
            for row in file_reader:
                attends_list.append(row)

        for attend in range(1, len(attends_list)):
            print(attends_list[attend][2])

            Attends.attends.create(
                uniq_id=attends_list[attend][1],
                group_id=Groups.groups.get(uniq_id=attends_list[attend][2]),
                user_id=Profile.objects.get(username=attends_list[attend][3]),
                online=True if attends_list[attend][6] == 'Да' else False,
                date_attend=attends_list[attend][7],
                start_time=attends_list[attend][8],
                end_time=attends_list[attend][9],
            )
            print(attend)
